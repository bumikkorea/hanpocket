#!/usr/bin/env python3
"""최종 병합: 카카오 + 네이버 + 미슐랭 + 블루리본 → 색상 엑셀"""
import json, re
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "/home/theredboat_ai/.openclaw/workspace/visa-app/src/data"
OUT = "/home/theredboat_ai/.openclaw/workspace/seoul_restaurants_FINAL.xlsx"

# 1. 기존 통합 데이터 로드
with open(f"{BASE}/seoul_restaurants_final.json", encoding="utf-8") as f:
    merged = json.load(f)
print(f"기존 통합: {len(merged)}개")

# 2. 미슐랭+블루리본 로드
with open("/tmp/michelin_blueribon.json", encoding="utf-8") as f:
    michelin = json.load(f)
print(f"미슐랭+블루리본: {len(michelin)}개")

# 3. 이름 정규화
def norm(name):
    if not name: return ""
    name = re.sub(r'<[^>]+>', '', name)
    name = re.sub(r'[·\-_()\s.]', '', name)
    return name.lower().strip()

# 4. 기존 데이터를 dict로
merged_dict = {}
for r in merged:
    key = norm(r["name"])
    merged_dict[key] = r

# 5. 미슐랭/블루리본 병합
AWARD_LABELS = {
    "michelin3": "⭐⭐⭐ 미슐랭 3스타",
    "michelin2": "⭐⭐ 미슐랭 2스타",
    "michelin1": "⭐ 미슐랭 1스타",
    "bib": "🍽️ 빕 구르망",
    "selected": "미슐랭 셀렉션",
    "blueribbon": "🔵 블루리본",
}

matched = 0
new_added = 0

for m in michelin:
    name_ko = m.get("name", {}).get("ko", "")
    key = norm(name_ko)
    
    award = AWARD_LABELS.get(m.get("award", ""), m.get("award", ""))
    
    if key in merged_dict:
        # 기존 항목에 미슐랭 정보 추가
        merged_dict[key]["michelin_award"] = award
        merged_dict[key]["michelin_url"] = m.get("michelinUrl", "")
        merged_dict[key]["catchtable_url"] = m.get("catchTableUrl", "")
        merged_dict[key]["images"] = m.get("images", [])
        merged_dict[key]["name_zh"] = m.get("name", {}).get("zh", "")
        merged_dict[key]["name_en"] = m.get("name", {}).get("en", "")
        if m.get("priceRange"):
            merged_dict[key]["price_range"] = m["priceRange"]
        matched += 1
    else:
        # 새 항목
        area_gu = m.get("area", {}).get("gu", "")
        # gu → area 매핑
        gu_area = {
            "강남구": "강남", "종로구": "종로", "중구": "명동",
            "마포구": "홍대", "용산구": "이태원", "성동구": "성수",
            "영등포구": "여의도", "송파구": "잠실",
        }
        area = gu_area.get(area_gu, "")
        
        new_entry = {
            "name": name_ko,
            "category": "음식점",
            "subcategory": "",
            "cuisine": m.get("cuisine", ""),
            "phone": "",
            "address": "",
            "address_old": "",
            "zipcode": "",
            "lat": m.get("lat", 0),
            "lng": m.get("lng", 0),
            "area": area,
            "review_count": 0,
            "rating": 0,
            "rating_count": 0,
            "popularity": "",
            "rating_grade": "",
            "photo_url": m.get("images", [""])[0] if m.get("images") else "",
            "parking": "",
            "wifi": "",
            "pet_friendly": "",
            "reservation": "",
            "delivery": "",
            "takeout": "",
            "disabled_access": "",
            "is_new_open": "",
            "brand": "",
            "kakao_url": "",
            "naver_url": "",
            "source": "michelin",
            "michelin_award": award,
            "michelin_url": m.get("michelinUrl", ""),
            "catchtable_url": m.get("catchTableUrl", ""),
            "images": m.get("images", []),
            "name_zh": m.get("name", {}).get("zh", ""),
            "name_en": m.get("name", {}).get("en", ""),
            "price_range": m.get("priceRange", 0),
        }
        merged_dict[key] = new_entry
        new_added += 1

print(f"  매칭: {matched}개 / 신규: {new_added}개")

# 미슐랭 없는 항목에도 빈 필드 추가
for key, r in merged_dict.items():
    if "michelin_award" not in r:
        r["michelin_award"] = ""
        r["michelin_url"] = ""
        r["catchtable_url"] = ""
        r["images"] = []
        r["name_zh"] = ""
        r["name_en"] = ""
        r["price_range"] = 0

# 정렬: 미슐랭 먼저, 그 다음 리뷰순
AWARD_ORDER = {
    "⭐⭐⭐ 미슐랭 3스타": 0, "⭐⭐ 미슐랭 2스타": 1, "⭐ 미슐랭 1스타": 2,
    "🍽️ 빕 구르망": 3, "미슐랭 셀렉션": 4, "🔵 블루리본": 5, "": 99
}

result = sorted(merged_dict.values(), key=lambda x: (
    AWARD_ORDER.get(x["michelin_award"], 99),
    -x.get("review_count", 0)
))

print(f"\n{'='*50}")
print(f"📊 최종 통합 결과")
print(f"{'='*50}")
print(f"총: {len(result)}개")

award_stats = Counter(r["michelin_award"] for r in result if r["michelin_award"])
for a, c in sorted(award_stats.items(), key=lambda x: AWARD_ORDER.get(x[0], 99)):
    print(f"  {a}: {c}개")
print(f"  일반: {sum(1 for r in result if not r['michelin_award'])}개")

# JSON 저장
with open(f"{BASE}/seoul_restaurants_FINAL.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print(f"\n📁 JSON 저장")

# ===== 엑셀 =====
wb = Workbook()
ws = wb.active
ws.title = "전체"

# 색상
YELLOW = PatternFill(start_color="FFF9C4", fill_type="solid")   # 카카오
GREEN = PatternFill(start_color="C8E6C9", fill_type="solid")    # 네이버
RED = PatternFill(start_color="FFCDD2", fill_type="solid")      # 미슐랭/블루리본
HEADER_K = PatternFill(start_color="F9A825", fill_type="solid")
HEADER_N = PatternFill(start_color="43A047", fill_type="solid")
HEADER_M = PatternFill(start_color="E53935", fill_type="solid")
HEADER = PatternFill(start_color="222222", fill_type="solid")
HFONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=9)
HFONT_B = Font(name="맑은 고딕", bold=True, color="000000", size=9)

cols = [
    # (이름, 폭, 출처) 
    ("순위", 5, "common"), ("이름", 22, "common"), ("중국어명", 10, "michelin"), ("영어명", 12, "michelin"),
    ("지역", 6, "common"), ("대분류", 8, "common"), ("중분류", 10, "common"), ("세분류", 10, "common"),
    ("미슐랭등급", 16, "michelin"), ("가격대", 5, "michelin"),
    ("리뷰수", 7, "kakao"), ("별점", 5, "kakao"), ("인기도", 10, "kakao"), ("평점등급", 8, "kakao"),
    ("사진URL", 40, "kakao"),
    ("주차", 4, "kakao"), ("와이파이", 5, "kakao"), ("반려동물", 5, "kakao"), ("예약", 4, "kakao"),
    ("배달", 4, "kakao"), ("포장", 4, "kakao"), ("장애인", 5, "kakao"),
    ("전화", 13, "common"), ("도로명주소", 35, "common"), ("위도", 10, "common"), ("경도", 10, "common"),
    ("카카오맵", 30, "kakao"), ("네이버", 30, "naver"),
    ("미슐랭링크", 30, "michelin"), ("캐치테이블", 30, "michelin"),
    ("출처", 6, "common")
]

FILL_MAP = {
    "kakao": (HEADER_K, YELLOW, HFONT_B),
    "naver": (HEADER_N, GREEN, HFONT),
    "michelin": (HEADER_M, RED, HFONT),
    "common": (HEADER, None, HFONT),
}

# 헤더
for col_idx, (h, w, src) in enumerate(cols, 1):
    c = ws.cell(row=1, column=col_idx, value=h)
    hf, _, font = FILL_MAP[src]
    c.fill = hf
    c.font = font
    c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# 범례
ws.insert_rows(2)
ws.cell(row=2, column=1, value="범례:").font = Font(bold=True, size=9)
ws.cell(row=2, column=2, value="🟡 노란색=카카오맵").fill = YELLOW
ws.cell(row=2, column=4, value="🟢 초록색=네이버").fill = GREEN
ws.cell(row=2, column=6, value="🔴 빨간색=미슐랭/블루리본").fill = RED
ws.cell(row=2, column=9, value="⬜ 흰색=공통")
for c in range(1, 11):
    ws.cell(row=2, column=c).font = Font(size=9)

PRICE_LABELS = {1: "₩", 2: "₩₩", 3: "₩₩₩", 4: "₩₩₩₩", 0: ""}
font = Font(name="맑은 고딕", size=9)

for i, r in enumerate(result):
    row = i + 3
    
    imgs = r.get("images", [])
    photo = r.get("photo_url", "") or (imgs[0] if imgs else "")
    
    vals = [
        i+1, r["name"], r.get("name_zh", ""), r.get("name_en", ""),
        r["area"], r["category"], r["subcategory"], r["cuisine"],
        r.get("michelin_award", ""), PRICE_LABELS.get(r.get("price_range", 0), ""),
        r["review_count"] or "", r["rating"] or "", r["popularity"], r["rating_grade"],
        photo,
        r["parking"], r["wifi"], r["pet_friendly"], r["reservation"],
        r["delivery"], r["takeout"], r["disabled_access"],
        r["phone"], r["address"], r["lat"], r["lng"],
        r["kakao_url"], r["naver_url"],
        r.get("michelin_url", ""), r.get("catchtable_url", ""),
        r["source"]
    ]
    
    for col_idx, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.font = font
        _, cell_fill, _ = FILL_MAP[cols[col_idx-1][2]]
        if cell_fill:
            cell.fill = cell_fill

ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(result)+2}"
ws.freeze_panes = "C3"

# 지역별 시트
ws2 = wb.create_sheet("지역별 TOP20")
by_area = defaultdict(list)
for r in result:
    if r["area"]:
        by_area[r["area"]].append(r)

AREAS = [
    "동대문", "잠실", "송파", "뚝섬", "건대", "강남", "홍대", "연남",
    "연희동", "신촌", "성수", "명동", "남산", "을지로", "한남", "이태원",
    "북촌", "인사동", "익선동", "삼청동", "여의도", "압구정", "청담",
    "서촌", "광화문", "종로", "코엑스", "삼성동"
]

row = 1
for area in AREAS:
    places = by_area.get(area, [])
    if not places: continue
    c = ws2.cell(row=row, column=1, value=f"📍 {area} ({len(places)}개)")
    c.font = Font(bold=True, size=11)
    c.fill = PatternFill(start_color="EEEEEE", fill_type="solid")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1
    for h_col, h in enumerate(["#","이름","세분류","미슐랭","인기도","별점","🅿️📶🐕","전화","주소"], 1):
        ws2.cell(row=row, column=h_col, value=h).font = Font(bold=True, size=9, color="888888")
    row += 1
    for i, r in enumerate(places[:20]):
        ws2.cell(row=row, column=1, value=i+1)
        ws2.cell(row=row, column=2, value=r["name"])
        ws2.cell(row=row, column=3, value=r["cuisine"] or r["subcategory"])
        award_cell = ws2.cell(row=row, column=4, value=r.get("michelin_award", ""))
        if r.get("michelin_award"):
            award_cell.fill = RED
        ws2.cell(row=row, column=5, value=r["popularity"])
        ws2.cell(row=row, column=6, value=f"{r['rating']}" if r["rating"] else "")
        amenities = []
        if r["parking"]=="Y": amenities.append("🅿️")
        if r["wifi"]=="Y": amenities.append("📶")
        if r["pet_friendly"]=="Y": amenities.append("🐕")
        ws2.cell(row=row, column=7, value="".join(amenities))
        ws2.cell(row=row, column=8, value=r["phone"])
        ws2.cell(row=row, column=9, value=r["address"])
        for c in range(1, 10):
            ws2.cell(row=row, column=c).font = Font(size=9)
        row += 1
    row += 1

for c, w in zip("ABCDEFGHI", [4, 22, 12, 16, 10, 5, 8, 13, 35]):
    ws2.column_dimensions[c].width = w

# 미슐랭 전용 시트
ws3 = wb.create_sheet("미슐랭 & 블루리본")
ws3.cell(row=1, column=1, value="⭐ 미슐랭 가이드 & 블루리본 전체 리스트").font = Font(bold=True, size=12)

m_headers = ["#", "등급", "이름", "중국어", "영어", "지역", "세분류", "가격대", "리뷰수", "별점", "미슐랭링크", "캐치테이블"]
for col, h in enumerate(m_headers, 1):
    c = ws3.cell(row=2, column=col, value=h)
    c.font = Font(bold=True, size=10)
    c.fill = RED

michelin_only = [r for r in result if r.get("michelin_award")]
for i, r in enumerate(michelin_only):
    row = i + 3
    ws3.cell(row=row, column=1, value=i+1)
    ws3.cell(row=row, column=2, value=r.get("michelin_award", ""))
    ws3.cell(row=row, column=3, value=r["name"])
    ws3.cell(row=row, column=4, value=r.get("name_zh", ""))
    ws3.cell(row=row, column=5, value=r.get("name_en", ""))
    ws3.cell(row=row, column=6, value=r["area"])
    ws3.cell(row=row, column=7, value=r["cuisine"] or r["subcategory"])
    ws3.cell(row=row, column=8, value=PRICE_LABELS.get(r.get("price_range", 0), ""))
    ws3.cell(row=row, column=9, value=r["review_count"] or "")
    ws3.cell(row=row, column=10, value=r["rating"] or "")
    ws3.cell(row=row, column=11, value=r.get("michelin_url", ""))
    ws3.cell(row=row, column=12, value=r.get("catchtable_url", ""))
    for c in range(1, 13):
        ws3.cell(row=row, column=c).font = Font(size=9)

for c, w in zip("ABCDEFGHIJKL", [4, 18, 20, 10, 15, 6, 12, 6, 7, 5, 35, 35]):
    ws3.column_dimensions[c].width = w

# 스키마 시트
ws4 = wb.create_sheet("스키마")
ws4.cell(row=1, column=1, value="필드명").font = Font(bold=True)
ws4.cell(row=1, column=2, value="설명").font = Font(bold=True)
ws4.cell(row=1, column=3, value="🟡카카오").font = Font(bold=True)
ws4.cell(row=1, column=3).fill = YELLOW
ws4.cell(row=1, column=4, value="🟢네이버").font = Font(bold=True)
ws4.cell(row=1, column=4).fill = GREEN
ws4.cell(row=1, column=5, value="🔴미슐랭").font = Font(bold=True)
ws4.cell(row=1, column=5).fill = RED
ws4.cell(row=1, column=6, value="타입").font = Font(bold=True)

schema = [
    ("name", "가게명", "✅", "✅", "✅", "text"),
    ("name_zh", "중국어명", "", "", "✅", "text"),
    ("name_en", "영어명", "", "", "✅", "text"),
    ("category", "대분류", "✅", "✅", "", "text"),
    ("subcategory", "중분류", "✅", "✅", "", "text"),
    ("cuisine", "세분류", "✅", "일부", "✅", "text"),
    ("michelin_award", "미슐랭등급", "", "", "✅", "text"),
    ("price_range", "가격대(1~4)", "", "", "✅", "int"),
    ("phone", "전화번호", "✅", "✅", "", "text"),
    ("address", "주소", "✅", "✅", "", "text"),
    ("lat/lng", "좌표", "✅", "✅", "✅", "float"),
    ("review_count", "리뷰수", "✅", "", "", "int"),
    ("rating", "별점", "✅", "", "", "float"),
    ("photo_url", "사진", "✅", "", "✅", "url"),
    ("parking", "주차(Y/N)", "✅", "", "", "Y/N"),
    ("wifi", "와이파이", "✅", "", "", "Y/N"),
    ("pet_friendly", "반려동물", "✅", "", "", "Y/N"),
    ("reservation", "예약", "✅", "", "", "Y/N"),
    ("kakao_url", "카카오맵", "✅", "", "", "url"),
    ("naver_url", "네이버", "", "✅", "", "url"),
    ("michelin_url", "미슐랭링크", "", "", "✅", "url"),
    ("catchtable_url", "캐치테이블", "", "", "✅", "url"),
]

for i, (f, d, k, n, m, t) in enumerate(schema, 2):
    ws4.cell(row=i, column=1, value=f).font = Font(size=10, name="Consolas")
    ws4.cell(row=i, column=2, value=d)
    c3 = ws4.cell(row=i, column=3, value=k)
    c4 = ws4.cell(row=i, column=4, value=n)
    c5 = ws4.cell(row=i, column=5, value=m)
    ws4.cell(row=i, column=6, value=t)
    if k: c3.fill = YELLOW
    if n: c4.fill = GREEN
    if m: c5.fill = RED

for c, w in zip("ABCDEF", [18, 15, 8, 8, 8, 6]):
    ws4.column_dimensions[c].width = w

wb.save(OUT)
print(f"\n📁 엑셀: {OUT}")
print("✅ 완료!")
