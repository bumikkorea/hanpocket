#!/usr/bin/env python3
"""최종 통합 엑셀 — 카카오(리뷰/별점/사진) + 네이버 병합"""
import json, re, time, urllib.request, urllib.parse
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = "/home/theredboat_ai/.openclaw/workspace/visa-app/src/data"
OUT = "/home/theredboat_ai/.openclaw/workspace/seoul_restaurants_final.xlsx"

# 카카오 인기순 데이터 (리뷰/별점 포함)
with open(f"{BASE}/seoul_restaurants_popular.json", encoding="utf-8") as f:
    kakao_data = json.load(f)

# 네이버 데이터
with open(f"{BASE}/naver_top_restaurants.json", encoding="utf-8") as f:
    naver_data = json.load(f)

print(f"카카오: {len(kakao_data)}개 / 네이버: {len(naver_data)}개")

# 카카오 검색 API에서 사진 URL 포함된 데이터 다시 수집
print("\n📸 사진 URL 포함 수집 중...")

AREAS = [
    "동대문", "잠실", "송파", "뚝섬", "건대", "강남", "홍대", "연남",
    "연희동", "신촌", "성수", "명동", "남산", "을지로", "한남", "이태원",
    "북촌", "인사동", "익선동", "삼청동", "여의도", "압구정", "청담",
    "서촌", "광화문", "종로", "코엑스", "삼성동"
]

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
    'Referer': 'https://map.kakao.com/',
}

photo_map = {}  # confirmid → img URL

for area in AREAS:
    query = urllib.parse.quote(f"{area} 맛집")
    for page in range(1, 6):
        url = f"https://search.map.kakao.com/mapsearch/map.daum?callback=j&q={query}&msFlag=A&sort=0&page={page}&cnt=15"
        req = urllib.request.Request(url, headers=HEADERS)
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                body = resp.read().decode()
                body = body[body.index('(') + 1 : body.rindex(')')]
                data = json.loads(body)
        except:
            break
        
        for p in data.get("place", []):
            pid = p.get("confirmid", "")
            img = p.get("img", "")
            if pid and img:
                photo_map[str(pid)] = img
            # 추가 정보도 저장
            if pid:
                photo_map[f"{pid}_parking"] = p.get("addinfo_parking", "")
                photo_map[f"{pid}_wifi"] = p.get("addinfo_wifi", "")
                photo_map[f"{pid}_pet"] = p.get("addinfo_pet", "")
                photo_map[f"{pid}_reservation"] = p.get("addinfo_appointment", "")
                photo_map[f"{pid}_delivery"] = p.get("addinfo_delivery", "")
                photo_map[f"{pid}_package"] = p.get("addinfo_package", "")
                photo_map[f"{pid}_newopen"] = str(p.get("is_new_open", ""))
        
        if data.get("meta", {}).get("is_end", True):
            if not data.get("place"):
                break
        time.sleep(0.15)
    time.sleep(0.1)

print(f"사진 확보: {sum(1 for k,v in photo_map.items() if not '_' in k and v)}개")

# 통합
def norm(name):
    if not name: return ""
    name = re.sub(r'<[^>]+>', '', name)
    name = re.sub(r'[·\-_()\s.]', '', name)
    return name.lower().strip()

merged = {}

for r in kakao_data:
    key = norm(r["name"])
    pid = str(r.get("id", ""))
    merged[key] = {
        "name": r["name"],
        "category": r.get("category", ""),
        "phone": r.get("phone", ""),
        "address": r.get("address", ""),
        "lat": r.get("lat", 0),
        "lng": r.get("lng", 0),
        "area": r.get("area", ""),
        "review_count": r.get("review_count", 0),
        "rating": r.get("rating", 0),
        "photo": photo_map.get(pid, ""),
        "parking": photo_map.get(f"{pid}_parking", ""),
        "wifi": photo_map.get(f"{pid}_wifi", ""),
        "pet": photo_map.get(f"{pid}_pet", ""),
        "reservation": photo_map.get(f"{pid}_reservation", ""),
        "kakao_url": r.get("kakao_url", ""),
        "naver_url": "",
        "source": "kakao",
    }

dup = 0
for r in naver_data:
    key = norm(r["name"])
    if key in merged:
        merged[key]["naver_url"] = r.get("link", "")
        merged[key]["source"] = "both"
        dup += 1
    else:
        merged[key] = {
            "name": re.sub(r'<[^>]+>', '', r["name"]),
            "category": r.get("category", ""),
            "phone": r.get("phone", ""),
            "address": r.get("address", ""),
            "lat": r.get("lat", 0),
            "lng": r.get("lng", 0),
            "area": "",
            "review_count": 0,
            "rating": 0,
            "photo": "",
            "parking": "", "wifi": "", "pet": "", "reservation": "",
            "kakao_url": "",
            "naver_url": r.get("link", ""),
            "source": "naver",
        }

result = sorted(merged.values(), key=lambda x: x["review_count"], reverse=True)
print(f"\n✅ 최종 통합: {len(result)}개 (중복매칭 {dup}개)")
print(f"  사진있는 곳: {sum(1 for r in result if r.get('photo'))}개")

# JSON 저장
with open(f"{BASE}/seoul_restaurants_final.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

# 인기순 등급
def popularity_grade(cnt):
    if cnt >= 1000: return "🔥🔥🔥 인기폭발"
    if cnt >= 500: return "🔥🔥 인기많음"
    if cnt >= 100: return "🔥 리뷰많음"
    if cnt > 0: return "리뷰있음"
    return ""

def rating_grade(r):
    if r >= 4.5: return "⭐ 최고"
    if r >= 4.0: return "⭐ 높음"
    if r >= 3.5: return "보통"
    if r > 0: return "낮음"
    return ""

# 엑셀 생성
wb = Workbook()
ws = wb.active
ws.title = "전체 인기순"

headers = ["순위", "이름", "지역", "카테고리", "인기도", "평점등급", "리뷰수", "별점",
           "주차", "와이파이", "반려동물", "예약", "전화번호", "주소", "사진URL", "카카오맵", "네이버"]

hfill = PatternFill(start_color="1F1F1F", fill_type="solid")
hfont = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hfont
    c.fill = hfill
    c.alignment = Alignment(horizontal='center')

for i, r in enumerate(result):
    row = i + 2
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=r["name"])
    ws.cell(row=row, column=3, value=r["area"])
    cat = r["category"].split(" > ")[-1] if " > " in r["category"] else r["category"]
    ws.cell(row=row, column=4, value=cat)
    ws.cell(row=row, column=5, value=popularity_grade(r["review_count"]))
    ws.cell(row=row, column=6, value=rating_grade(r["rating"]))
    ws.cell(row=row, column=7, value=r["review_count"])
    ws.cell(row=row, column=8, value=r["rating"])
    ws.cell(row=row, column=9, value="✅" if r.get("parking") == "Y" else "")
    ws.cell(row=row, column=10, value="✅" if r.get("wifi") == "Y" else "")
    ws.cell(row=row, column=11, value="✅" if r.get("pet") == "Y" else "")
    ws.cell(row=row, column=12, value="✅" if r.get("reservation") == "Y" else "")
    ws.cell(row=row, column=13, value=r.get("phone", ""))
    ws.cell(row=row, column=14, value=r["address"])
    ws.cell(row=row, column=15, value=r.get("photo", ""))
    ws.cell(row=row, column=16, value=r["kakao_url"])
    ws.cell(row=row, column=17, value=r["naver_url"])
    
    for col in range(1, 18):
        ws.cell(row=row, column=col).font = Font(name="맑은 고딕", size=9)

widths = [5, 22, 6, 18, 12, 10, 7, 5, 4, 5, 5, 4, 13, 35, 40, 30, 30]
from openpyxl.utils import get_column_letter
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.auto_filter.ref = f"A1:Q{len(result)+1}"

# 지역별 시트
ws2 = wb.create_sheet("지역별 TOP 20")
by_area = defaultdict(list)
for r in result:
    if r["area"]:
        by_area[r["area"]].append(r)

row = 1
for area in AREAS:
    places = by_area.get(area, [])
    if not places: continue
    
    c = ws2.cell(row=row, column=1, value=f"📍 {area} ({len(places)}개)")
    c.font = Font(name="맑은 고딕", bold=True, size=11)
    c.fill = PatternFill(start_color="F0F0F0", fill_type="solid")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    
    sub = ["#", "이름", "카테고리", "인기도", "평점", "주차/와이파이", "전화", "주소"]
    for col, h in enumerate(sub, 1):
        ws2.cell(row=row, column=col, value=h).font = Font(name="맑은 고딕", bold=True, size=9, color="888888")
    row += 1
    
    for i, r in enumerate(places[:20]):
        ws2.cell(row=row, column=1, value=i+1)
        ws2.cell(row=row, column=2, value=r["name"])
        cat = r["category"].split(" > ")[-1] if " > " in r["category"] else r["category"]
        ws2.cell(row=row, column=3, value=cat)
        ws2.cell(row=row, column=4, value=popularity_grade(r["review_count"]))
        ws2.cell(row=row, column=5, value=rating_grade(r["rating"]))
        amenities = []
        if r.get("parking") == "Y": amenities.append("🅿️")
        if r.get("wifi") == "Y": amenities.append("📶")
        ws2.cell(row=row, column=6, value=" ".join(amenities))
        ws2.cell(row=row, column=7, value=r.get("phone", ""))
        ws2.cell(row=row, column=8, value=r["address"])
        for col in range(1, 9):
            ws2.cell(row=row, column=col).font = Font(name="맑은 고딕", size=9)
        row += 1
    row += 1

ws2.column_dimensions['A'].width = 4
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 8
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 13
ws2.column_dimensions['H'].width = 35

# 통계 시트
ws3 = wb.create_sheet("통계")
ws3.cell(row=1, column=1, value="📊 서울 맛집 데이터 통계").font = Font(bold=True, size=14)

stats = [
    ("총 맛집 수", len(result)),
    ("카카오 단독", sum(1 for r in result if r["source"]=="kakao")),
    ("네이버 단독", sum(1 for r in result if r["source"]=="naver")),
    ("양쪽 확인", sum(1 for r in result if r["source"]=="both")),
    ("사진 보유", sum(1 for r in result if r.get("photo"))),
    ("리뷰 보유", sum(1 for r in result if r["review_count"] > 0)),
    ("별점 보유", sum(1 for r in result if r["rating"] > 0)),
    ("좌표 보유", sum(1 for r in result if r["lat"] and r["lng"])),
    ("인기폭발(1000+)", sum(1 for r in result if r["review_count"] >= 1000)),
    ("인기많음(500+)", sum(1 for r in result if 500 <= r["review_count"] < 1000)),
    ("주차가능", sum(1 for r in result if r.get("parking") == "Y")),
    ("와이파이", sum(1 for r in result if r.get("wifi") == "Y")),
]

for i, (label, val) in enumerate(stats, 3):
    ws3.cell(row=i, column=1, value=label).font = Font(size=11)
    ws3.cell(row=i, column=2, value=val).font = Font(size=11, bold=True)

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 10

wb.save(OUT)
print(f"\n📁 엑셀 저장: {OUT}")
