#!/usr/bin/env python3
"""맛집 데이터 엑셀 변환"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/home/theredboat_ai/.openclaw/workspace/visa-app/src/data"
OUT = "/home/theredboat_ai/.openclaw/workspace/seoul_restaurants_popular.xlsx"

with open(f"{BASE}/seoul_restaurants_popular.json", encoding="utf-8") as f:
    data = json.load(f)

wb = Workbook()

# --- Sheet 1: 전체 인기순 ---
ws = wb.active
ws.title = "전체 인기순"

headers = ["순위", "이름", "지역", "카테고리", "리뷰수", "별점", "전화번호", "주소", "위도", "경도", "카카오맵"]
header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
thin = Side(style='thin', color='DDDDDD')
border = Border(bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for i, r in enumerate(data):
    row = i + 2
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=r["name"])
    ws.cell(row=row, column=3, value=r["area"])
    ws.cell(row=row, column=4, value=r["category"])
    ws.cell(row=row, column=5, value=r["review_count"])
    ws.cell(row=row, column=6, value=r["rating"])
    ws.cell(row=row, column=7, value=r.get("phone", ""))
    ws.cell(row=row, column=8, value=r["address"])
    ws.cell(row=row, column=9, value=r["lat"])
    ws.cell(row=row, column=10, value=r["lng"])
    ws.cell(row=row, column=11, value=r["kakao_url"])
    
    for col in range(1, 12):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).font = Font(name="맑은 고딕", size=10)

# 컬럼 폭
widths = [5, 25, 7, 25, 8, 6, 15, 40, 10, 10, 35]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 필터
ws.auto_filter.ref = f"A1:K{len(data)+1}"

# --- Sheet 2: 지역별 시트 ---
from collections import defaultdict
by_area = defaultdict(list)
for r in data:
    by_area[r["area"]].append(r)

AREAS = [
    "동대문", "잠실", "송파", "뚝섬", "건대", "강남", "홍대", "연남",
    "연희동", "신촌", "성수", "명동", "남산", "을지로", "한남", "이태원",
    "북촌", "인사동", "익선동", "삼청동", "여의도", "압구정", "청담",
    "서촌", "광화문", "종로", "코엑스", "삼성동"
]

ws2 = wb.create_sheet("지역별 TOP")
row = 1
for area in AREAS:
    places = by_area.get(area, [])
    if not places:
        continue
    
    # 지역 헤더
    cell = ws2.cell(row=row, column=1, value=f"📍 {area} ({len(places)}개)")
    cell.font = Font(name="맑은 고딕", bold=True, size=12)
    cell.fill = PatternFill(start_color="F0F0F0", fill_type="solid")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    
    # 서브 헤더
    sub_headers = ["#", "이름", "카테고리", "리뷰수", "별점", "전화번호", "주소"]
    for col, h in enumerate(sub_headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = Font(name="맑은 고딕", bold=True, size=10, color="666666")
    row += 1
    
    for i, r in enumerate(places[:20]):  # 지역당 TOP 20
        ws2.cell(row=row, column=1, value=i+1)
        ws2.cell(row=row, column=2, value=r["name"])
        cat_short = r["category"].split(" > ")[-1] if r["category"] else ""
        ws2.cell(row=row, column=3, value=cat_short)
        ws2.cell(row=row, column=4, value=r["review_count"])
        ws2.cell(row=row, column=5, value=r["rating"])
        ws2.cell(row=row, column=6, value=r.get("phone", ""))
        ws2.cell(row=row, column=7, value=r["address"])
        for col in range(1, 8):
            ws2.cell(row=row, column=col).font = Font(name="맑은 고딕", size=10)
        row += 1
    
    row += 1  # 빈 줄

ws2.column_dimensions['A'].width = 4
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 6
ws2.column_dimensions['F'].width = 15
ws2.column_dimensions['G'].width = 40

wb.save(OUT)
print(f"✅ 엑셀 저장: {OUT}")
print(f"   전체 인기순: {len(data)}개")
print(f"   지역별 TOP: {len(AREAS)}개 지역")
