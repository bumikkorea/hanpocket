#!/usr/bin/env python3
"""카카오 + 네이버 표준화 통합 — 빠르게, 사진/편의정보 포함"""
import json, re, time, urllib.request, urllib.parse
from collections import Counter, defaultdict

BASE = "/home/theredboat_ai/.openclaw/workspace/visa-app/src/data"

# 1. 카카오 검색 API에서 풀데이터 재수집 (사진/편의정보 포함)
print("📸 카카오 풀데이터 수집 중...")

AREAS = [
    "동대문", "잠실", "송파", "뚝섬", "건대", "강남", "홍대", "연남",
    "연희동", "신촌", "성수", "명동", "남산", "을지로", "한남", "이태원",
    "북촌", "인사동", "익선동", "삼청동", "여의도", "압구정", "청담",
    "서촌", "광화문", "종로", "코엑스", "삼성동"
]

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
    'Referer': 'https://map.kakao.com/',
}

kakao_full = {}  # id → full data
seen = set()

for area in AREAS:
    query = urllib.parse.quote(f"{area} 맛집")
    for page in range(1, 6):
        url = f"https://search.map.kakao.com/mapsearch/map.daum?callback=j&q={query}&msFlag=A&sort=0&page={page}&cnt=15"
        req = urllib.request.Request(url, headers=HEADERS)
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                body = resp.read().decode()
                body = body[body.index('(') + 1 : body.rindex(')')]
                data = json.loads(body)
        except:
            break
        
        places = data.get("place", [])
        if not places:
            break
            
        for p in places:
            pid = str(p.get("confirmid", ""))
            if pid in seen:
                continue
            seen.add(pid)
            
            cat1 = p.get("cate_name_depth1", "")
            if cat1 and cat1 not in ["음식점", "카페,디저트", ""]:
                continue
            
            kakao_full[pid] = {
                "kakao_id": pid,
                "name": p.get("name", ""),
                "category_depth1": p.get("cate_name_depth1", ""),
                "category_depth2": p.get("cate_name_depth2", ""),
                "category_depth3": p.get("cate_name_depth3", ""),
                "phone": p.get("tel", ""),
                "address": p.get("new_address", "") or p.get("address", ""),
                "address_old": p.get("address", ""),
                "zipcode": p.get("new_zipcode", ""),
                "lat": float(p.get("lat", 0) or 0),
                "lng": float(p.get("lon", 0) or 0),
                "review_count": int(p.get("reviewCount", 0) or 0),
                "rating_avg": float(p.get("rating_average", 0) or 0),
                "rating_count": int(p.get("rating_count", 0) or 0),
                "photo_url": p.get("img", ""),
                "parking": p.get("addinfo_parking", ""),
                "wifi": p.get("addinfo_wifi", ""),
                "pet_friendly": p.get("addinfo_pet", ""),
                "reservation": p.get("addinfo_appointment", ""),
                "delivery": p.get("addinfo_delivery", ""),
                "takeout": p.get("addinfo_package", ""),
                "disabled_access": p.get("addinfo_fordisabled", ""),
                "is_new_open": str(p.get("is_new_open", "")),
                "brand": p.get("brandName", ""),
                "kakao_url": f"https://place.map.kakao.com/{pid}",
                "area": area,
            }
        
        time.sleep(0.15)
    time.sleep(0.1)

print(f"  카카오 수집: {len(kakao_full)}개")

# 2. 네이버 데이터 로드
with open(f"{BASE}/naver_top_restaurants.json", encoding="utf-8") as f:
    naver_data = json.load(f)
print(f"  네이버 로드: {len(naver_data)}개")

# 3. 표준 스키마 정의 & 통합
def norm(name):
    if not name: return ""
    name = re.sub(r'<[^>]+>', '', name)
    name = re.sub(r'[·\-_()\s.]', '', name)
    return name.lower().strip()

# 표준 필드
STANDARD_FIELDS = [
    "name",                # 가게명
    "category",            # 대분류 (음식점/카페)
    "subcategory",         # 중분류 (한식/양식/중식)
    "cuisine",             # 세분류 (냉면/파스타/초밥)
    "phone",               # 전화번호
    "address",             # 도로명주소
    "address_old",         # 지번주소
    "zipcode",             # 우편번호
    "lat",                 # 위도
    "lng",                 # 경도
    "area",                # 관광지역
    "review_count",        # 리뷰수 (카카오)
    "rating",              # 별점 (카카오)
    "rating_count",        # 별점참여수 (카카오)
    "popularity",          # 인기도 등급
    "rating_grade",        # 평점 등급
    "photo_url",           # 대표사진 URL (카카오)
    "parking",             # 주차 Y/N (카카오)
    "wifi",                # 와이파이 Y/N (카카오)
    "pet_friendly",        # 반려동물 Y/N (카카오)
    "reservation",         # 예약가능 Y/N (카카오)
    "delivery",            # 배달 Y/N (카카오)
    "takeout",             # 포장 Y/N (카카오)
    "disabled_access",     # 장애인편의 Y/N (카카오)
    "is_new_open",         # 신규오픈 (카카오)
    "brand",               # 브랜드명 (카카오)
    "kakao_url",           # 카카오맵 링크
    "naver_url",           # 네이버 링크
    "source",              # 데이터 출처
]

def popularity_grade(cnt):
    if cnt >= 1000: return "인기폭발"
    if cnt >= 500: return "인기많음"
    if cnt >= 100: return "리뷰많음"
    if cnt > 0: return "리뷰있음"
    return ""

def rating_grade(r):
    if r >= 4.5: return "최고"
    if r >= 4.0: return "높음"
    if r >= 3.5: return "보통"
    if r > 0: return "낮음"
    return ""

# 지역 감지
AREA_MAP = {
    "동대문": ["동대문", "신설동", "제기동"],
    "잠실": ["잠실", "신천동"], "송파": ["송파", "방이동", "문정동"],
    "뚝섬": ["뚝섬", "성수동", "서울숲"], "건대": ["화양동", "자양동"],
    "강남": ["강남", "역삼", "서초", "논현", "신사동"],
    "홍대": ["서교동", "상수동", "합정"], "연남": ["연남동"],
    "연희동": ["연희동"], "신촌": ["신촌", "창천동"],
    "성수": ["성수"], "명동": ["명동", "충무로"],
    "남산": ["남산", "후암동"], "을지로": ["을지로", "초동"],
    "한남": ["한남"], "이태원": ["이태원", "녹사평"],
    "북촌": ["가회동", "계동"], "인사동": ["인사동", "관훈동"],
    "익선동": ["익선동"], "삼청동": ["삼청"],
    "여의도": ["여의도"], "압구정": ["압구정"],
    "청담": ["청담"], "서촌": ["통인동", "효자동", "체부동"],
    "광화문": ["광화문", "세종"], "종로": ["종로", "관철동"],
    "코엑스": ["코엑스", "봉은사"], "삼성동": ["삼성동"],
}

def detect_area(addr):
    if not addr: return ""
    for area, kws in AREA_MAP.items():
        if any(kw in addr for kw in kws):
            return area
    return ""

# 카카오 → 표준
merged = {}
for pid, k in kakao_full.items():
    key = norm(k["name"])
    merged[key] = {
        "name": k["name"],
        "category": k["category_depth1"],
        "subcategory": k["category_depth2"],
        "cuisine": k["category_depth3"],
        "phone": k["phone"],
        "address": k["address"],
        "address_old": k["address_old"],
        "zipcode": k["zipcode"],
        "lat": k["lat"],
        "lng": k["lng"],
        "area": k["area"],
        "review_count": k["review_count"],
        "rating": k["rating_avg"],
        "rating_count": k["rating_count"],
        "popularity": popularity_grade(k["review_count"]),
        "rating_grade": rating_grade(k["rating_avg"]),
        "photo_url": k["photo_url"],
        "parking": k["parking"],
        "wifi": k["wifi"],
        "pet_friendly": k["pet_friendly"],
        "reservation": k["reservation"],
        "delivery": k["delivery"],
        "takeout": k["takeout"],
        "disabled_access": k["disabled_access"],
        "is_new_open": k["is_new_open"],
        "brand": k["brand"],
        "kakao_url": k["kakao_url"],
        "naver_url": "",
        "source": "kakao",
    }

# 네이버 병합
dup = 0
for r in naver_data:
    key = norm(r["name"])
    clean_name = re.sub(r'<[^>]+>', '', r["name"])
    
    if key in merged:
        merged[key]["naver_url"] = r.get("link", "")
        merged[key]["source"] = "both"
        if not merged[key]["phone"] and r.get("phone"):
            merged[key]["phone"] = r["phone"]
        dup += 1
    else:
        addr = r.get("address", "")
        cat_parts = r.get("category", "").split(">")
        merged[key] = {
            "name": clean_name,
            "category": cat_parts[0].strip() if cat_parts else "",
            "subcategory": cat_parts[1].strip() if len(cat_parts) > 1 else "",
            "cuisine": cat_parts[2].strip() if len(cat_parts) > 2 else "",
            "phone": r.get("phone", ""),
            "address": r.get("roadAddress", addr),
            "address_old": addr,
            "zipcode": "",          # 네이버 미제공
            "lat": r.get("lat", 0),
            "lng": r.get("lng", 0),
            "area": detect_area(addr),
            "review_count": 0,      # 네이버 미제공
            "rating": 0,            # 네이버 미제공
            "rating_count": 0,      # 네이버 미제공
            "popularity": "",       # 네이버 미제공
            "rating_grade": "",     # 네이버 미제공
            "photo_url": "",        # 네이버 미제공
            "parking": "",          # 네이버 미제공
            "wifi": "",             # 네이버 미제공
            "pet_friendly": "",     # 네이버 미제공
            "reservation": "",      # 네이버 미제공
            "delivery": "",         # 네이버 미제공
            "takeout": "",          # 네이버 미제공
            "disabled_access": "",  # 네이버 미제공
            "is_new_open": "",      # 네이버 미제공
            "brand": "",            # 네이버 미제공
            "kakao_url": "",
            "naver_url": r.get("link", ""),
            "source": "naver",
        }

result = sorted(merged.values(), key=lambda x: x["review_count"], reverse=True)

# 통계
print(f"\n{'='*50}")
print(f"📊 최종 통합 결과")
print(f"{'='*50}")
print(f"총 맛집: {len(result)}개")
print(f"  카카오 단독: {sum(1 for r in result if r['source']=='kakao')}개")
print(f"  네이버 단독: {sum(1 for r in result if r['source']=='naver')}개")
print(f"  양쪽 확인: {sum(1 for r in result if r['source']=='both')}개")
print(f"  중복 매칭: {dup}개")
print(f"\n데이터 커버리지:")
print(f"  좌표: {sum(1 for r in result if r['lat'] and r['lng'])}개")
print(f"  사진: {sum(1 for r in result if r['photo_url'])}개")
print(f"  리뷰: {sum(1 for r in result if r['review_count']>0)}개")
print(f"  별점: {sum(1 for r in result if r['rating']>0)}개")
print(f"  주차: {sum(1 for r in result if r['parking']=='Y')}개")
print(f"  와이파이: {sum(1 for r in result if r['wifi']=='Y')}개")
print(f"  예약: {sum(1 for r in result if r['reservation']=='Y')}개")
print(f"  반려동물: {sum(1 for r in result if r['pet_friendly']=='Y')}개")
print(f"  지역분류: {sum(1 for r in result if r['area'])}개")

# JSON 저장
out_json = f"{BASE}/seoul_restaurants_final.json"
with open(out_json, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print(f"\n📁 JSON: {out_json}")

# 엑셀 저장
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_XLS = "/home/theredboat_ai/.openclaw/workspace/seoul_restaurants_final.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "전체 (인기순)"

excel_cols = [
    ("순위", 5), ("이름", 22), ("지역", 6), ("대분류", 8), ("중분류", 10), ("세분류", 10),
    ("인기도", 10), ("평점등급", 8), ("리뷰수", 7), ("별점", 5), ("별점수", 6),
    ("사진URL", 40), ("주차", 4), ("와이파이", 5), ("반려동물", 5), ("예약", 4),
    ("배달", 4), ("포장", 4), ("장애인", 5), ("신규", 4), ("브랜드", 10),
    ("전화", 13), ("도로명주소", 35), ("지번주소", 35), ("우편번호", 7),
    ("위도", 10), ("경도", 10), ("카카오맵", 30), ("네이버", 30), ("출처", 6)
]

hfill = PatternFill(start_color="222222", fill_type="solid")
hfont = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=9)

for col, (h, w) in enumerate(excel_cols, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hfont
    c.fill = hfill
    c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(col)].width = w

font = Font(name="맑은 고딕", size=9)
for i, r in enumerate(result):
    row = i + 2
    vals = [
        i+1, r["name"], r["area"], r["category"], r["subcategory"], r["cuisine"],
        r["popularity"], r["rating_grade"], r["review_count"] or "", r["rating"] or "", r["rating_count"] or "",
        r["photo_url"], r["parking"], r["wifi"], r["pet_friendly"], r["reservation"],
        r["delivery"], r["takeout"], r["disabled_access"], r["is_new_open"], r["brand"],
        r["phone"], r["address"], r["address_old"], r["zipcode"],
        r["lat"], r["lng"], r["kakao_url"], r["naver_url"], r["source"]
    ]
    for col, v in enumerate(vals, 1):
        ws.cell(row=row, column=col, value=v).font = font

ws.auto_filter.ref = f"A1:{get_column_letter(len(excel_cols))}{len(result)+1}"
ws.freeze_panes = "C2"

# 지역별 시트
ws2 = wb.create_sheet("지역별 TOP20")
by_area = defaultdict(list)
for r in result:
    if r["area"]:
        by_area[r["area"]].append(r)

row = 1
for area in AREAS:
    places = by_area.get(area, [])
    if not places: continue
    c = ws2.cell(row=row, column=1, value=f"📍 {area} ({len(places)}개)")
    c.font = Font(bold=True, size=11)
    c.fill = PatternFill(start_color="EEEEEE", fill_type="solid")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    for h_col, h in enumerate(["#","이름","세분류","인기도","별점","편의","전화","주소"], 1):
        ws2.cell(row=row, column=h_col, value=h).font = Font(bold=True, size=9, color="888888")
    row += 1
    for i, r in enumerate(places[:20]):
        ws2.cell(row=row, column=1, value=i+1)
        ws2.cell(row=row, column=2, value=r["name"])
        ws2.cell(row=row, column=3, value=r["cuisine"] or r["subcategory"])
        ws2.cell(row=row, column=4, value=r["popularity"])
        ws2.cell(row=row, column=5, value=f"{r['rating']}" if r["rating"] else "")
        amenities = []
        if r["parking"]=="Y": amenities.append("🅿️")
        if r["wifi"]=="Y": amenities.append("📶")
        if r["pet_friendly"]=="Y": amenities.append("🐕")
        if r["reservation"]=="Y": amenities.append("📅")
        ws2.cell(row=row, column=6, value="".join(amenities))
        ws2.cell(row=row, column=7, value=r["phone"])
        ws2.cell(row=row, column=8, value=r["address"])
        for c in range(1,9):
            ws2.cell(row=row, column=c).font = Font(size=9)
        row += 1
    row += 1

for c, w in zip("ABCDEFGH", [4,22,12,10,5,8,13,35]):
    ws2.column_dimensions[c].width = w

# 스키마 시트
ws3 = wb.create_sheet("스키마 (데이터 사전)")
ws3.cell(row=1, column=1, value="필드명").font = Font(bold=True)
ws3.cell(row=1, column=2, value="설명").font = Font(bold=True)
ws3.cell(row=1, column=3, value="카카오").font = Font(bold=True)
ws3.cell(row=1, column=4, value="네이버").font = Font(bold=True)
ws3.cell(row=1, column=5, value="타입").font = Font(bold=True)

schema = [
    ("name", "가게명", "✅", "✅", "text"),
    ("category", "대분류", "✅", "✅", "text"),
    ("subcategory", "중분류", "✅", "✅", "text"),
    ("cuisine", "세분류 (요리종류)", "✅", "일부", "text"),
    ("phone", "전화번호", "✅", "✅", "text"),
    ("address", "도로명주소", "✅", "✅", "text"),
    ("address_old", "지번주소", "✅", "✅", "text"),
    ("zipcode", "우편번호", "✅", "", "text"),
    ("lat/lng", "위도/경도", "✅", "✅", "float"),
    ("area", "관광지역 (28개)", "✅", "주소기반추론", "text"),
    ("review_count", "리뷰수", "✅", "", "int"),
    ("rating", "별점 (5점만점)", "✅", "", "float"),
    ("rating_count", "별점참여수", "✅", "", "int"),
    ("popularity", "인기도 등급", "리뷰수 기반", "", "text"),
    ("rating_grade", "평점 등급", "별점 기반", "", "text"),
    ("photo_url", "대표사진 URL", "✅", "", "url"),
    ("parking", "주차가능 (Y/N)", "✅", "", "Y/N"),
    ("wifi", "와이파이 (Y/N)", "✅", "", "Y/N"),
    ("pet_friendly", "반려동물 (Y/N)", "✅", "", "Y/N"),
    ("reservation", "예약가능 (Y/N)", "✅", "", "Y/N"),
    ("delivery", "배달 (Y/N)", "✅", "", "Y/N"),
    ("takeout", "포장 (Y/N)", "✅", "", "Y/N"),
    ("disabled_access", "장애인편의 (Y/N)", "✅", "", "Y/N"),
    ("is_new_open", "신규오픈", "✅", "", "bool"),
    ("brand", "브랜드명", "✅", "", "text"),
    ("kakao_url", "카카오맵 링크", "✅", "", "url"),
    ("naver_url", "네이버 링크", "", "✅", "url"),
    ("source", "데이터 출처", "kakao/both", "naver/both", "text"),
]

for i, (field, desc, kakao, naver, dtype) in enumerate(schema, 2):
    ws3.cell(row=i, column=1, value=field).font = Font(size=10, name="Consolas")
    ws3.cell(row=i, column=2, value=desc).font = Font(size=10)
    ws3.cell(row=i, column=3, value=kakao).font = Font(size=10)
    ws3.cell(row=i, column=4, value=naver).font = Font(size=10)
    ws3.cell(row=i, column=5, value=dtype).font = Font(size=10)

for c, w in zip("ABCDE", [18, 20, 12, 12, 8]):
    ws3.column_dimensions[c].width = w

wb.save(OUT_XLS)
print(f"📁 엑셀: {OUT_XLS}")
print("\n✅ 완료!")
