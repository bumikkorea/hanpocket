#!/usr/bin/env python3
"""엑셀에 출처별 색상 적용 — 카카오:노란색 / 네이버:초록색 / 양쪽:흰색"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

BASE = "/home/theredboat_ai/.openclaw/workspace/visa-app/src/data"
OUT = "/home/theredboat_ai/.openclaw/workspace/seoul_restaurants_final.xlsx"

with open(f"{BASE}/seoul_restaurants_final.json", encoding="utf-8") as f:
    data = json.load(f)

print(f"로드: {len(data)}개")

# 색상 정의
YELLOW = PatternFill(start_color="FFF9C4", fill_type="solid")  # 카카오 (연노랑)
GREEN = PatternFill(start_color="C8E6C9", fill_type="solid")   # 네이버 (연초록)
WHITE = PatternFill(start_color="FFFFFF", fill_type="solid")    # 양쪽
HEADER = PatternFill(start_color="1F1F1F", fill_type="solid")
HFONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=9)

# 카카오 전용 필드 (노란색)
KAKAO_COLS = {"리뷰수", "별점", "별점수", "인기도", "평점등급", "사진URL", 
              "주차", "와이파이", "반려동물", "예약", "배달", "포장", 
              "장애인", "신규", "브랜드", "우편번호", "카카오맵"}

# 네이버 전용 필드 (초록색)
NAVER_COLS = {"네이버"}

# 양쪽 공통 필드 (흰색)
COMMON_COLS = {"순위", "이름", "지역", "대분류", "중분류", "세분류",
               "전화", "도로명주소", "지번주소", "위도", "경도", "출처"}

wb = Workbook()
ws = wb.active
ws.title = "전체 (인기순)"

cols = [
    ("순위", 5), ("이름", 22), ("지역", 6), 
    ("대분류", 8), ("중분류", 10), ("세분류", 10),
    ("리뷰수", 7), ("별점", 5), ("별점수", 6), ("인기도", 10), ("평점등급", 8),
    ("사진URL", 40),
    ("주차", 4), ("와이파이", 5), ("반려동물", 5), ("예약", 4),
    ("배달", 4), ("포장", 4), ("장애인", 5), ("신규", 4), ("브랜드", 10),
    ("전화", 13), ("도로명주소", 35), ("지번주소", 35), ("우편번호", 7),
    ("위도", 10), ("경도", 10), ("카카오맵", 30), ("네이버", 30), ("출처", 6)
]

# 헤더 — 색상으로 출처 표시
for col_idx, (h, w) in enumerate(cols, 1):
    c = ws.cell(row=1, column=col_idx, value=h)
    c.font = HFONT
    if h in KAKAO_COLS:
        c.fill = PatternFill(start_color="F9A825", fill_type="solid")  # 진한 노랑
    elif h in NAVER_COLS:
        c.fill = PatternFill(start_color="43A047", fill_type="solid")  # 진한 초록
    else:
        c.fill = HEADER
    c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# 범례 행
ws.insert_rows(2)
ws.cell(row=2, column=1, value="범례:").font = Font(bold=True, size=9)
ws.cell(row=2, column=2, value="🟡 노란색 = 카카오맵 데이터").font = Font(size=9)
ws.cell(row=2, column=2).fill = YELLOW
ws.cell(row=2, column=5, value="🟢 초록색 = 네이버 데이터").font = Font(size=9)
ws.cell(row=2, column=5).fill = GREEN
ws.cell(row=2, column=8, value="⬜ 흰색 = 공통 데이터").font = Font(size=9)

font = Font(name="맑은 고딕", size=9)

for i, r in enumerate(data):
    row = i + 3  # 헤더 + 범례 다음
    
    vals = [
        i+1, r["name"], r["area"],
        r["category"], r["subcategory"], r["cuisine"],
        r["review_count"] or "", r["rating"] or "", r["rating_count"] or "",
        r["popularity"], r["rating_grade"],
        r["photo_url"],
        r["parking"], r["wifi"], r["pet_friendly"], r["reservation"],
        r["delivery"], r["takeout"], r["disabled_access"], r["is_new_open"], r["brand"],
        r["phone"], r["address"], r["address_old"], r["zipcode"],
        r["lat"], r["lng"], r["kakao_url"], r["naver_url"], r["source"]
    ]
    
    for col_idx, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.font = font
        
        h = cols[col_idx-1][0]
        if h in KAKAO_COLS:
            cell.fill = YELLOW
        elif h in NAVER_COLS:
            cell.fill = GREEN
        # 공통은 흰색 (기본)

ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data)+2}"
ws.freeze_panes = "C3"

# 지역별 시트
AREAS = [
    "동대문", "잠실", "송파", "뚝섬", "건대", "강남", "홍대", "연남",
    "연희동", "신촌", "성수", "명동", "남산", "을지로", "한남", "이태원",
    "북촌", "인사동", "익선동", "삼청동", "여의도", "압구정", "청담",
    "서촌", "광화문", "종로", "코엑스", "삼성동"
]

ws2 = wb.create_sheet("지역별 TOP20")
by_area = defaultdict(list)
for r in data:
    if r["area"]:
        by_area[r["area"]].append(r)

row = 1
for area in AREAS:
    places = by_area.get(area, [])
    if not places: continue
    c = ws2.cell(row=row, column=1, value=f"📍 {area} ({len(places)}개)")
    c.font = Font(bold=True, size=11)
    c.fill = PatternFill(start_color="EEEEEE", fill_type="solid")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1
    
    sub_h = ["#","이름","세분류","인기도","별점","🅿️","📶","🐕","전화","주소"]
    for h_col, h in enumerate(sub_h, 1):
        ws2.cell(row=row, column=h_col, value=h).font = Font(bold=True, size=9, color="888888")
    row += 1
    
    for i, r in enumerate(places[:20]):
        ws2.cell(row=row, column=1, value=i+1)
        ws2.cell(row=row, column=2, value=r["name"])
        ws2.cell(row=row, column=3, value=r["cuisine"] or r["subcategory"])
        ws2.cell(row=row, column=4, value=r["popularity"]).fill = YELLOW
        ws2.cell(row=row, column=5, value=f"{r['rating']}" if r["rating"] else "").fill = YELLOW
        ws2.cell(row=row, column=6, value="✅" if r["parking"]=="Y" else "").fill = YELLOW
        ws2.cell(row=row, column=7, value="✅" if r["wifi"]=="Y" else "").fill = YELLOW
        ws2.cell(row=row, column=8, value="✅" if r["pet_friendly"]=="Y" else "").fill = YELLOW
        ws2.cell(row=row, column=9, value=r["phone"])
        ws2.cell(row=row, column=10, value=r["address"])
        for c in range(1, 11):
            ws2.cell(row=row, column=c).font = Font(size=9)
        row += 1
    row += 1

for c, w in zip("ABCDEFGHIJ", [4, 22, 12, 10, 5, 4, 4, 4, 13, 35]):
    ws2.column_dimensions[c].width = w

# 스키마 시트
ws3 = wb.create_sheet("스키마")
ws3.cell(row=1, column=1, value="필드명").font = Font(bold=True)
ws3.cell(row=1, column=2, value="설명").font = Font(bold=True)
ws3.cell(row=1, column=3, value="카카오맵").font = Font(bold=True)
ws3.cell(row=1, column=3).fill = PatternFill(start_color="F9A825", fill_type="solid")
ws3.cell(row=1, column=4, value="네이버").font = Font(bold=True)
ws3.cell(row=1, column=4).fill = PatternFill(start_color="43A047", fill_type="solid")
ws3.cell(row=1, column=5, value="타입").font = Font(bold=True)

schema = [
    ("name", "가게명", "✅", "✅", "text"),
    ("category", "대분류", "✅", "✅", "text"),
    ("subcategory", "중분류", "✅", "✅", "text"),
    ("cuisine", "세분류", "✅", "일부", "text"),
    ("phone", "전화번호", "✅", "✅", "text"),
    ("address", "도로명주소", "✅", "✅", "text"),
    ("address_old", "지번주소", "✅", "✅", "text"),
    ("zipcode", "우편번호", "✅", "", "text"),
    ("lat/lng", "좌표", "✅", "✅", "float"),
    ("area", "관광지역", "✅", "주소추론", "text"),
    ("review_count", "리뷰수", "✅", "", "int"),
    ("rating", "별점", "✅", "", "float"),
    ("rating_count", "별점참여수", "✅", "", "int"),
    ("popularity", "인기도 등급", "✅기반", "", "text"),
    ("rating_grade", "평점 등급", "✅기반", "", "text"),
    ("photo_url", "대표사진", "✅", "", "url"),
    ("parking", "주차", "✅", "", "Y/N"),
    ("wifi", "와이파이", "✅", "", "Y/N"),
    ("pet_friendly", "반려동물", "✅", "", "Y/N"),
    ("reservation", "예약", "✅", "", "Y/N"),
    ("delivery", "배달", "✅", "", "Y/N"),
    ("takeout", "포장", "✅", "", "Y/N"),
    ("disabled_access", "장애인편의", "✅", "", "Y/N"),
    ("is_new_open", "신규오픈", "✅", "", "bool"),
    ("brand", "브랜드명", "✅", "", "text"),
    ("kakao_url", "카카오맵링크", "✅", "", "url"),
    ("naver_url", "네이버링크", "", "✅", "url"),
]

for i, (f, d, k, n, t) in enumerate(schema, 2):
    ws3.cell(row=i, column=1, value=f).font = Font(size=10, name="Consolas")
    ws3.cell(row=i, column=2, value=d)
    c3 = ws3.cell(row=i, column=3, value=k)
    c4 = ws3.cell(row=i, column=4, value=n)
    ws3.cell(row=i, column=5, value=t)
    if k: c3.fill = YELLOW
    if n: c4.fill = GREEN

for c, w in zip("ABCDE", [18, 15, 10, 10, 8]):
    ws3.column_dimensions[c].width = w

wb.save(OUT)
print(f"📁 {OUT}")
print("✅ 색상 적용 완료! 🟡카카오 🟢네이버 ⬜공통")
